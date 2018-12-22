import csv

from openpyxl import load_workbook


headers = ['year', 'property_category', 'property_type', 'distinct_incidents',
           'money_seized', 'lbs_seized', 'msrp', 'quantity_seized']

infile_name = 'dhs-forfeitures-2014-2017-raw.xlsx'
outfile_name = 'dhs-forfeitures-2014-2017-parsed.csv'

wb = load_workbook(filename=infile_name)


def extract_data(sheet_name):
    '''given the name of a sheet, return a actual data minus cruft'''

    # grab the year from the sheet name
    year = int(sheet_name.split()[-1].strip())

    # get a handle to the data in the sheet
    data = wb[sheet_name]

    # need to chop off some cruft above and below the actual data
    # so loop over the data and put down start/stop markers
    start = 0
    stop = 0

    # enumerate the rows of data
    for i, row in enumerate(data.values):

        # if the opening value is 'PROPERTY CATEGORY'
        # we're at the header row, so we want to start on the next row
        if str(row[0]) == 'PROPERTY CATEGORY':
            start = i+1

        # If we're at the row with the page number, that's our stop point
        if str(row[3]).startswith('Page '):
            stop = i

    # return a dict with the year and the correctly sliced data
    return {
        'year': year,
        'data': list(data.values)[start:stop]
    }


def parse_data(year=None, data=None):
    '''given a dict with year and correctly sliced data,
    parse into a clean list of dictionaries'''

    # start off with no product category
    product_category = None

    # a list to hold the data that will be returned
    parsed_data = []

    # loop over the data
    for row in data:

        # turn tuple into a list so that later we can insert the year
        row = list(row)

        # if there's a value in the first position, it's a new category
        if row[0]:
            # set new value for product category and keep on truckin'
            product_category = row[0].strip().upper()
            continue

        # if values are missing in the first two positions
        # that's a summary row that we want to skip
        if not row[0] and not row[1]:
            continue

        # fill in product category
        if not row[0]:
            row[0] = product_category

        # slide the year into the first position
        row.insert(0, year)

        # turn it all into a dict and append to the big list of parsed data
        clean_dict = dict(zip(headers, row))
        parsed_data.append(clean_dict)

    return parsed_data


with open(outfile_name, 'w', newline='', encoding='utf-8') as outfile:

    # set up writer object and write headers to file
    writer = csv.DictWriter(outfile, fieldnames=headers)
    writer.writeheader()

    # loop over the sheets
    for sheet in wb.sheetnames:

        # extract the data
        sheet_data = extract_data(sheet)

        # parse the data
        parsed_data = parse_data(**sheet_data)

        # write the rows to file
        writer.writerows(parsed_data)
